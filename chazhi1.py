
import numpy as np #导入numpy模块，并赋予numpy一个别名：np（numpy模块提供了很多数据处理函数，用来处理相同类型，固定长度的元素）
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d#从scipy.interpolate模块中引入一维插值函数interp1d
from scipy.interpolate import griddata

import os
import openpyxl
from openpyxl import Workbook


workbook = openpyxl.load_workbook('C:/Users/13183/PycharmProjects/pythonProject/chuli.xlsx')
sheetsource = workbook['原始温度数据']

for i in sheetsource.iter_rows(min_row=3, min_col=1, max_col=37):
    sheetname = str(i[0].value)
    sheetname = sheetname.split(' ')[0]
    sheetname = sheetname + '(17张图)'
    workbook.create_sheet(sheetname)
    sheet = workbook[sheetname]

    avg = float((i[1].value+i[2].value)/2)
    for j in range(3,71):
        sheet.cell(row=j, column=3).value = avg
        sheet.cell(row=j, column=5).value = avg
        sheet.cell(row=j, column=7).value = avg
        sheet.cell(row=j, column=9).value = avg

    c7 = i[34:37]
    y = [cell.value for cell in c7]
    x= np.array([5198,3368,1538])
    y= np.array(y)
    f = interp1d(x, y, kind = 'quadratic') # 得到插值函数
    x1=list(range(3500, 5100+100,100))#3500 5100，半径方向 0-4600深度方向
    y1=f(x1)
    for j in range(3,71):
        k=int((j-3)/4)
        print(y1[k])
        sheet.cell(row=j,column=4).value = y1[k]

    c7 = i[31:34]
    y = [cell.value for cell in c7]
    x = np.array([5192,3432,1672])
    y = np.array(y)
    print(y)
    f = interp1d(x, y, kind='quadratic')  # 得到插值函数
    x1 = list(range(3500, 5100 + 100, 100))  # 3500 5100，半径方向 0-4600深度方向
    y1 = f(x1)

    for j in range(3, 71):
        k = int((j - 3) / 4)
        print(y1[k])
        sheet.cell(row=j, column=6).value = y1[k]

    c7 = i[28:31]
    y = [cell.value for cell in c7]
    x = np.array([5211,3711,2211])
    y = np.array(y)
    f = interp1d(x, y, kind='quadratic')  # 得到插值函数
    x1 = list(range(3500, 5100 + 100, 100))  # 3500 5100，半径方向 0-4600深度方向
    y1 = f(x1)

    for j in range(3, 71):
        k = int((j - 3) / 4)
        print(y1[k])
        sheet.cell(row=j, column=8).value = y1[k]

    c7 = i[22:28]
    y = [cell.value for cell in c7]
    x= np.array([14990,12640,10290,7940,5590,3420])
    y= np.array(y)
    f = interp1d(x, y, kind = 'quadratic') # 得到插值函数
    x1=list(range(3500, 5100+100,100))#3500 5100，半径方向 0-4600深度方向
    y1=f(x1)
    k=int(0)
    for j in range(6,71,4):
        print(y1[k])
        sheet.cell(row=j,column=10).value = y1[k]
        k=k+1

    c7 = i[16:22]
    y = [cell.value for cell in c7]
    x = np.array([15100,12709,10318,7927,5536,3145])
    y = np.array(y)
    f = interp1d(x, y, kind='quadratic')  # 得到插值函数
    x1 = list(range(3500, 5100 + 100, 100))  # 3500 5100，半径方向 0-4600深度方向
    y1 = f(x1)
    k = int(0)
    for j in range(5, 71, 4):
        print(y1[k])
        sheet.cell(row=j, column=10).value = y1[k]
        k = k + 1

    c7 = i[10:16]
    y = [cell.value for cell in c7]
    x = np.array([15000,12400,9800,7200,4600,2000])
    y = np.array(y)
    f = interp1d(x, y, kind='quadratic')  # 得到插值函数
    x1 = list(range(3500, 5100 + 100, 100))  # 3500 5100，半径方向 0-4600深度方向
    y1 = f(x1)
    k = int(0)
    for j in range(4, 71, 4):
        print(y1[k])
        sheet.cell(row=j, column=10).value = y1[k]
        k = k + 1

    c7 = i[4:10]
    y = [cell.value for cell in c7]
    x = np.array([15000,12400,9800,7200,4600,2000])
    y = np.array(y)
    f = interp1d(x, y, kind='quadratic')  # 得到插值函数
    x1 = list(range(3500, 5100 + 100, 100))  # 3500 5100，半径方向 0-4600深度方向
    y1 = f(x1)
    k = int(0)
    for j in range(3, 71, 4):
        print(y1[k])
        sheet.cell(row=j, column=10).value = y1[k]
        k = k + 1

x=np.array([0,675,1350,2025,2700,3300,3900,4600])
y=[]
n=0

for row_num, row in enumerate(sheet.iter_rows(min_row=3, max_row=70, min_col=3, max_col=10), start=3):
    y = [cell.value for cell in row]
    f = interp1d(x, y, kind='quadratic')
    x1 = list(range(x.min(), x.max() + 100, 100))
    y1 = f(x1)
    for idx, val in enumerate(y1):
        sheet.cell(row=row_num, column=12+idx, value=val)

workbook.save("C:/Users/13183/PycharmProjects/pythonProject/chuli.xlsx")